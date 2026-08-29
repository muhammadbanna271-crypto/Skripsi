"""
Generator file Excel (.xlsx) untuk itinerary/rundown perjalanan.

Dibangun dari data ``TripPlanningService`` (bukan hasil karangan LLM):
itinerary + estimasi biaya + data destinasi. Output berupa bytes xlsx
yang siap dikirim sebagai HTTP attachment.

Struktur:
    Sheet 1 "Rundown"          -> tabel per hari (waktu, kegiatan, destinasi)
    Sheet 2 "Ringkasan Biaya"  -> komponen biaya + total
    Sheet 3 "Destinasi"        -> daftar destinasi + status + Google Maps link
"""

import io
from datetime import timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from apps.gis.models import TouristDestination, format_idr
from apps.gis.services.trip_planning import TripPlanningService


WEEKDAY_NAMES_ID = [
    "Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu",
]
MONTH_NAMES_ID = [
    "Januari", "Februari", "Maret", "April", "Mei", "Juni",
    "Juli", "Agustus", "September", "Oktober", "November", "Desember",
]

# Gaya sel (warna senada tema TRIP: hijau hutan).
HEADER_FILL = PatternFill("solid", fgColor="166534")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14, color="14532D")
INFO_FONT = Font(size=10, color="445248")
DAY_FILL = PatternFill("solid", fgColor="E7F0E3")
DAY_FONT = Font(bold=True, color="14532D")
TOTAL_FILL = PatternFill("solid", fgColor="FDE68A")
TOTAL_FONT = Font(bold=True, color="14532D")
LINK_FONT = Font(color="0563C1", underline="single")
NOTE_FONT = Font(size=9, italic=True, color="6B7280")

_THIN = Side(style="thin", color="D9E6DA")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="top", wrap_text=True)

RUPIAH_FORMAT = '"Rp"#,##0'


def _day_label(day):
    """Format tanggal Indonesia: 'Sabtu, 27 Agustus 2026'."""
    if day is None:
        return "tanggal belum ditentukan"
    return (
        f"{WEEKDAY_NAMES_ID[day.weekday()]}, {day.day} "
        f"{MONTH_NAMES_ID[day.month - 1]} {day.year}"
    )


def _format_duration(minutes):
    """Ubah menit jadi label ringkas: '2 jam' / '1 jam 30 mnt' / '30 mnt'."""
    if minutes is None:
        return "—"
    minutes = int(minutes)
    hours, mins = divmod(minutes, 60)
    if hours and mins:
        return f"{hours} jam {mins} mnt"
    if hours:
        return f"{hours} jam"
    return f"{mins} mnt"


class ItineraryExcelService:
    """Bangun workbook xlsx itinerary yang profesional."""

    @classmethod
    def build(cls, params, destination_ids=None):
        itinerary = TripPlanningService.build_itinerary(
            params, destination_ids=destination_ids
        )
        budget = TripPlanningService.estimate_budget(
            params, destination_ids=destination_ids
        )
        destinations = cls._destinations_by_id(itinerary)
        return cls._render(itinerary, budget, destinations, params)

    @classmethod
    def _destinations_by_id(cls, itinerary):
        """Ambil objek destinasi yang benar-benar masuk itinerary (urut)."""
        ids = []
        for day in itinerary.get("days", []):
            for item in day.get("items", []):
                if item.get("type") == "destination" and item.get("destination_id"):
                    if item["destination_id"] not in ids:
                        ids.append(item["destination_id"])

        by_id = {
            dest.id: dest
            for dest in TouristDestination.objects
            .filter(pk__in=ids)
            .select_related("village", "district")
            .prefetch_related(
                "categories",
                "cuisine_types",
                "wahanas",
                "wahanas__bundles",
                "bundles",
                "bundles__wahanas",
                "parking_fees",
            )
        }
        return [by_id[i] for i in ids if i in by_id]

    # ---------------------------------------------------------
    # RENDER
    # ---------------------------------------------------------

    @classmethod
    def _render(cls, itinerary, budget, destinations, params):
        wb = Workbook()

        cls._sheet_rundown(wb.active, itinerary, budget, destinations, params)
        cls._sheet_budget(wb.create_sheet("Ringkasan Biaya"), budget, params)
        cls._sheet_destinations(wb.create_sheet("Destinasi"), destinations)
        cls._sheet_detail_prices(wb.create_sheet("Detail Harga"), destinations)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def _title(sheet, text, columns):
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)
        cell = sheet.cell(row=1, column=1, value=text)
        cell.font = TITLE_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        sheet.row_dimensions[1].height = 26

    @staticmethod
    def _info_line(sheet, row, text, columns):
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=columns)
        cell = sheet.cell(row=row, column=1, value=text)
        cell.font = INFO_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")

    @classmethod
    def _write_header(cls, sheet, row, columns):
        for col, header in enumerate(columns, start=1):
            cell = sheet.cell(row=row, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
            cell.border = BORDER

    # ---------------------------------------------------------
    # SHEET 1 — RUNDOWN
    # ---------------------------------------------------------

    @classmethod
    def _sheet_rundown(cls, sheet, itinerary, budget, destinations, params):
        sheet.title = "Rundown"
        columns = ["No", "Waktu", "Kegiatan", "Destinasi", "Durasi", "Estimasi Tiket", "Catatan"]
        n_cols = len(columns)

        cls._title(sheet, "RUNDOWN PERJALANAN WISATA KOTA BATU", n_cols)

        start_date = TripPlanningService._parse_date(params.get("start_date"))
        duration = itinerary.get("duration_days", 1)
        travelers = params.get("traveler_count") or 1
        ticket_total = budget.get("known_total", 0)
        info = (
            f"Tanggal: {_day_label(start_date)}   •   "
            f"Durasi: {duration} hari   •   "
            f"Jumlah Wisatawan: {travelers} orang   •   "
            f"Estimasi Tiket & Parkir: {format_idr(ticket_total) or 'belum tersedia'}"
        )
        cls._info_line(sheet, 2, info, n_cols)

        cls._write_header(sheet, 3, columns)

        trip_dates = cls._trip_dates(start_date, duration)

        row = 4
        no = 0
        for index, day in enumerate(itinerary.get("days", [])):
            day_label = _day_label(trip_dates[index] if index < len(trip_dates) else None)
            sheet.merge_cells(
                start_row=row, start_column=1, end_row=row, end_column=n_cols
            )
            cell = sheet.cell(row=row, column=1, value=f"Hari {day.get('day', index + 1)} — {day_label}")
            cell.fill = DAY_FILL
            cell.font = DAY_FONT
            for col in range(1, n_cols + 1):
                sheet.cell(row=row, column=col).fill = DAY_FILL
                sheet.cell(row=row, column=col).border = BORDER
            row += 1

            for item in day.get("items", []):
                no += 1
                is_lunch = item.get("type") == "lunch"
                activity = "Makan siang" if is_lunch else "Wisata"
                price = "" if is_lunch else (item.get("price_display") or "Belum tersedia")

                values = [
                    no,
                    item.get("time", ""),
                    activity,
                    item.get("name", ""),
                    _format_duration(item.get("duration_minutes")),
                    price,
                    item.get("note", "") or "",
                ]
                for col, value in enumerate(values, start=1):
                    cell = sheet.cell(row=row, column=col, value=value)
                    cell.border = BORDER
                    cell.alignment = RIGHT if col == 2 else LEFT
                row += 1

        for unscheduled in itinerary.get("unscheduled", []):
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
            cell = sheet.cell(
                row=row, column=1,
                value=f"Tidak dijadwalkan: {unscheduled.get('name')} — {unscheduled.get('note', '')}",
            )
            cell.font = NOTE_FONT
            row += 1

        widths = [6, 20, 14, 30, 14, 20, 30]
        for i, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + i)].width = width

        sheet.freeze_panes = "A4"

    @staticmethod
    def _trip_dates(start_date, duration):
        if start_date is None:
            return []
        return [start_date + timedelta(days=i) for i in range(duration)]

    # ---------------------------------------------------------
    # SHEET 2 — RINGKASAN BIAYA
    # ---------------------------------------------------------

    @classmethod
    def _sheet_budget(cls, sheet, budget, params):
        columns = ["Komponen", "Estimasi"]
        cls._write_header(sheet, 1, columns)

        ticket_total = sum(
            item.get("ticket") or 0 for item in budget.get("known_breakdown", [])
        )
        parking_total = sum(
            item.get("parking") or 0 for item in budget.get("known_breakdown", [])
        )
        transport = budget.get("transport_estimate") or {}
        transport_cost = transport.get("fuel_cost")

        rows = [
            ("Tiket wisata", ticket_total, True),
            ("Parkir", parking_total, True),
            ("Transportasi (bahan bakar, estimasi)", transport_cost, True),
            ("Makan", None, False),
            ("Lain-lain", None, False),
        ]

        row = 2
        for label, value, _is_numeric in rows:
            sheet.cell(row=row, column=1, value=label).border = BORDER
            value_cell = sheet.cell(row=row, column=2)
            value_cell.border = BORDER
            if value is None:
                value_cell.value = "belum diestimasi"
                value_cell.font = NOTE_FONT
            else:
                value_cell.value = value
                value_cell.number_format = RUPIAH_FORMAT
            value_cell.alignment = RIGHT
            row += 1

        known_total = ticket_total + parking_total + (transport_cost or 0)
        total_cell = sheet.cell(row=row, column=1, value="Total (diketahui)")
        total_cell.font = TOTAL_FONT
        total_cell.fill = TOTAL_FILL
        total_cell.border = BORDER
        total_value = sheet.cell(row=row, column=2, value=known_total)
        total_value.number_format = RUPIAH_FORMAT
        total_value.font = TOTAL_FONT
        total_value.fill = TOTAL_FILL
        total_value.border = BORDER
        total_value.alignment = RIGHT
        row += 2

        note = (
            "Estimasi hanya mencakup tiket & parkir yang tersedia di database "
            "serta perkiraan bahan bakar (garis lurus). Makan dan lain-lain "
            "belum diestimasi otomatis."
        )
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        note_cell = sheet.cell(row=row, column=1, value=note)
        note_cell.font = NOTE_FONT
        note_cell.alignment = LEFT

        sheet.column_dimensions["A"].width = 44
        sheet.column_dimensions["B"].width = 22
        sheet.freeze_panes = "A2"

    # ---------------------------------------------------------
    # SHEET 3 — DESTINASI
    # ---------------------------------------------------------

    @classmethod
    def _sheet_destinations(cls, sheet, destinations):
        columns = [
            "No", "Destinasi", "Status", "Jam Operasional",
            "Harga", "Wahana", "Bundle", "Google Maps",
        ]
        cls._write_header(sheet, 1, columns)

        for index, dest in enumerate(destinations, start=1):
            row = index + 1
            values = [
                index,
                dest.name,
                dest.status_label,
                dest.operating_hours_display,
                dest.price_display,
                dest.ride_prices_display or "—",
                dest.bundle_prices_display or "—",
                "Buka Google Maps",
            ]
            for col, value in enumerate(values, start=1):
                cell = sheet.cell(row=row, column=col, value=value)
                cell.border = BORDER
                cell.alignment = CENTER if col in (1, 3) else LEFT

            link_cell = sheet.cell(row=row, column=8)
            link_cell.hyperlink = dest.google_maps_url()
            link_cell.font = LINK_FONT
            link_cell.alignment = LEFT

        widths = [6, 28, 12, 18, 22, 24, 26, 18]
        for i, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + i)].width = width
        sheet.freeze_panes = "A2"

    # ---------------------------------------------------------
    # SHEET 4 — DETAIL HARGA (HTM / Wahana / Bundle)
    # ---------------------------------------------------------

    @classmethod
    def _sheet_detail_prices(cls, sheet, destinations):
        columns = ["Jenis", "Nama", "Harga / Status", "Keterangan"]
        n_cols = len(columns)

        cls._title(sheet, "DETAIL HARGA (HTM, WAHANA, BUNDLE)", n_cols)
        cls._write_header(sheet, 2, columns)

        row = 3
        for dest in destinations:
            sheet.merge_cells(
                start_row=row, start_column=1, end_row=row, end_column=n_cols
            )
            cell = sheet.cell(row=row, column=1, value=dest.name)
            cell.fill = DAY_FILL
            cell.font = DAY_FONT
            for col in range(1, n_cols + 1):
                sheet.cell(row=row, column=col).fill = DAY_FILL
                sheet.cell(row=row, column=col).border = BORDER
            row += 1

            rows = [("HTM", dest.name, dest.price_display, "Tiket masuk")]

            for w in dest.wahanas.all():
                if not w.is_active:
                    continue
                ket = w.get_pricing_type_display()
                if w.pricing_type == "INCLUDED_IN_PACKAGE":
                    bnames = [b.name for b in w.bundles.all()]
                    if bnames:
                        ket += " — Paket: " + ", ".join(bnames)
                rows.append(("Wahana", w.name, w.price_display, ket))

            for b in dest.bundles.all():
                if not b.is_active:
                    continue
                notes = []
                if b.includes_entry_ticket:
                    notes.append("Termasuk HTM")
                if b.rides_display:
                    notes.append(f"Isi: {b.rides_display}")
                rows.append(
                    ("Bundle", b.name, b.price_display, "; ".join(notes) or "—")
                )

            for f in dest.parking_fees.all():
                if not f.is_active:
                    continue
                rows.append(
                    ("Parkir", f.vehicle_type, f.price_display, "Biaya parkir")
                )

            for jenis, nama, harga, ket in rows:
                sheet.cell(row=row, column=1, value=jenis).border = BORDER
                sheet.cell(row=row, column=2, value=nama).border = BORDER
                sheet.cell(row=row, column=3, value=harga).border = BORDER
                sheet.cell(row=row, column=4, value=ket).border = BORDER
                row += 1

            row += 1  # baris kosong antar destinasi

        widths = [12, 30, 22, 40]
        for i, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + i)].width = width
        sheet.freeze_panes = "A3"
