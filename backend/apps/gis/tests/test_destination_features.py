import json
import os
import tempfile
from datetime import date, time
from io import BytesIO

from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from django.core.management import call_command
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from apps.gis.forms import TouristDestinationForm
from apps.gis.models import TicketBundle, TouristDestination, Wahana
from apps.gis.services.itinerary_excel import ItineraryExcelService
from apps.gis.services.price_source import PriceSourceService
from apps.gis.services.trip_planning import TripPlanningService
from apps.master.models import District, Village


class DestinationFeatureTest(TestCase):
    """Status, harga fleksibel, 24 jam, Google Maps, Excel, update harga."""

    def setUp(self):
        self.district = District.objects.create(code="T1", name="Test Kecamatan")
        self.village = Village.objects.create(
            code="T1V1",
            name="Desa Test",
            district=self.district,
            latitude=-7.8695,
            longitude=112.5236,
        )

    def _destination(self, name, **kwargs):
        defaults = {
            "name": name,
            "village": self.village,
            "latitude": -7.8695,
            "longitude": 112.5236,
            "estimated_duration_minutes": 120,
        }
        defaults.update(kwargs)
        return TouristDestination.objects.create(**defaults)

    # ---------------- STATUS ----------------

    def test_status_property_and_reason(self):
        dest = self._destination("Coban X")
        self.assertEqual(dest.status, "active")
        self.assertIsNotNone(dest.status_updated_at)

        dest.is_active = False
        dest.status_reason = "Sedang renovasi"
        dest.save()
        dest.refresh_from_db()

        self.assertEqual(dest.status, "inactive")
        self.assertEqual(dest.status_label, "Nonaktif")
        self.assertEqual(dest.status_reason, "Sedang renovasi")
        self.assertIsNotNone(dest.status_updated_at)

    def test_inactive_excluded_from_search(self):
        active = self._destination("Destinasi Aktif")
        inactive = self._destination("Destinasi Nonaktif", is_active=False)

        result = TripPlanningService.search_destinations({})
        names = [r["name"] for r in result["results"]]
        self.assertIn(active.name, names)
        self.assertNotIn(inactive.name, names)

    def test_get_details_finds_inactive_with_reason(self):
        # Detail tetap bisa ditemukan supaya chatbot bisa menjelaskan
        # alasan nonaktif; hanya rekomendasi/itinerary yang mengecualikan.
        dest = self._destination(
            "Coban Renovasi", is_active=False, status_reason="Sedang renovasi"
        )
        details = TripPlanningService.get_destination_details(
            destination_id=dest.id
        )
        self.assertTrue(details["found"])
        self.assertEqual(details["status"], "inactive")
        self.assertEqual(details["status_reason"], "Sedang renovasi")

    def test_inactive_excluded_from_itinerary(self):
        active = self._destination("Destinasi Aktif")
        inactive = self._destination("Destinasi Nonaktif", is_active=False)

        itinerary = TripPlanningService.build_itinerary(
            {"duration_days": 1, "destination_ids": [active.id, inactive.id]}
        )
        scheduled = [
            item["name"]
            for day in itinerary["days"]
            for item in day["items"]
            if item["type"] == "destination"
        ]
        self.assertIn(active.name, scheduled)
        self.assertNotIn(inactive.name, scheduled)

    # ---------------- HARGA (gratis vs belum diketahui) ----------------

    def test_free_destination_is_known_zero_not_unknown(self):
        dest = self._destination("Gratis", is_free=True)
        self.assertEqual(dest.ticket_price_for(), 0)

        budget = TripPlanningService.estimate_budget(
            {"destination_ids": [dest.id]}
        )
        self.assertEqual(budget["known_total"], 0)
        self.assertEqual(len(budget["known_breakdown"]), 1)
        self.assertEqual(len(budget["unknown_items"]), 0)

    def test_unknown_destination_is_not_free(self):
        dest = self._destination("Belum Tahu", ticket_type="unknown")
        budget = TripPlanningService.estimate_budget(
            {"destination_ids": [dest.id]}
        )
        self.assertEqual(budget["known_total"], 0)
        self.assertEqual(len(budget["known_breakdown"]), 0)
        self.assertEqual(len(budget["unknown_items"]), 1)

    def test_category_price_min_and_display(self):
        dest = self._destination(
            "Per Kategori",
            ticket_type="category",
            category_prices=[
                {"category": "Dewasa", "price": 20000},
                {"category": "Anak-anak", "price": 10000},
            ],
        )
        self.assertEqual(dest.min_category_price, 10000)
        self.assertIn("Dewasa", dest.price_display)
        self.assertIn("Anak-anak", dest.price_display)

    # ---------------- PARKIR ----------------

    def test_free_parking_display_and_cost(self):
        dest = self._destination("Parkir Gratis", is_free_parking=True)
        self.assertEqual(dest.parking_cost_int, 0)
        self.assertEqual(dest.parking_display, "Gratis")

    def test_parking_independent_from_free_entry(self):
        # Masuk gratis bukan berarti parkir gratis.
        dest = self._destination(
            "Masuk Gratis Parkir Bayar", is_free=True, parking_cost=5000
        )
        self.assertEqual(dest.parking_cost_int, 5000)
        self.assertEqual(dest.parking_display, "Rp5.000")

    def test_parking_surfaced_in_details(self):
        dest = self._destination("Parkir 5rb", parking_cost=5000)
        details = TripPlanningService.get_destination_details(
            destination_id=dest.id
        )
        self.assertEqual(details["parking_display"], "Rp5.000")
        self.assertFalse(details["is_free_parking"])

    # ---------------- WAHANA ----------------

    def test_ride_prices_display(self):
        dest = self._destination("Wahana Bayar")
        Wahana.objects.create(
            destination=dest, name="Flying Fox", pricing_type="INDEPENDENT_PRICE", price=25000
        )
        Wahana.objects.create(
            destination=dest, name="ATV", pricing_type="INDEPENDENT_PRICE", price=50000
        )
        self.assertIn("Flying Fox Rp25.000", dest.ride_prices_display)
        self.assertIn("ATV Rp50.000", dest.ride_prices_display)

        details = TripPlanningService.get_destination_details(
            destination_id=dest.id
        )
        self.assertIn("Flying Fox Rp25.000", details["ride_prices_display"])
        self.assertEqual(len(details["wahanas"]), 2)

    def test_bundle_prices_display(self):
        dest = self._destination("Bundle")
        ff = Wahana.objects.create(
            destination=dest, name="Flying Fox", pricing_type="INDEPENDENT_PRICE", price=25000
        )
        bundle = TicketBundle.objects.create(
            destination=dest,
            name="Paket Hemat",
            includes_entry_ticket=True,
            price=50000,
        )
        bundle.wahanas.add(ff)
        self.assertIn("Paket Hemat", dest.bundle_prices_display)
        self.assertIn("HTM", dest.bundle_prices_display)
        self.assertIn("Rp50.000", dest.bundle_prices_display)

        details = TripPlanningService.get_destination_details(
            destination_id=dest.id
        )
        self.assertEqual(len(details["bundles"]), 1)

    # ---------------- ELEVASI & SUHU ----------------

    def test_elevation_temperature_serialized(self):
        dest = self._destination(
            "Elevasi Suhu",
            elevation_meters=900,
            elevation_source="DEMNAS",
            temperature_c=19.5,
            temperature_source="WorldClim",
        )
        details = TripPlanningService.get_destination_details(
            destination_id=dest.id
        )
        self.assertEqual(details["elevation_meters"], 900)
        self.assertEqual(details["elevation_source"], "DEMNAS")
        self.assertEqual(details["temperature_c"], 19.5)
        self.assertEqual(details["temperature_source"], "WorldClim")

    def test_import_destination_climate(self):
        dest = self._destination("Alun-Alun Import")
        path = None
        try:
            with tempfile.NamedTemporaryFile(
                mode="w", suffix=".csv", delete=False,
                newline="", encoding="utf-8",
            ) as fh:
                fh.write(
                    "name,elevation_meters,elevation_source,"
                    "temperature_c,temperature_source,temperature_date\n"
                )
                fh.write(
                    "Alun-Alun Import,900,DEMNAS,19.5,WorldClim,2026-01-01\n"
                )
                path = fh.name

            call_command("import_destination_climate", path)
            dest.refresh_from_db()
            self.assertEqual(dest.elevation_meters, 900)
            self.assertEqual(dest.elevation_source, "DEMNAS")
            self.assertEqual(dest.temperature_c, 19.5)
            self.assertEqual(dest.temperature_source, "WorldClim")
            self.assertEqual(dest.temperature_date, date(2026, 1, 1))
        finally:
            if path and os.path.exists(path):
                os.remove(path)

    # ---------------- JAM OPERASIONAL ----------------

    def test_operating_hours_24h(self):
        dest = self._destination("Buka 24 Jam", is_open_24_hours=True)
        self.assertEqual(dest.operating_hours_display, "Buka 24 jam")

    def test_operating_hours_normal(self):
        dest = self._destination(
            "Normal",
            opening_time=time(8, 0),
            closing_time=time(17, 0),
        )
        self.assertEqual(dest.operating_hours_display, "08:00–17:00")

    # ---------------- GOOGLE MAPS ----------------

    def test_google_maps_url_uses_name_query(self):
        dest = self._destination("Alun-Alun", latitude=-7.8695, longitude=112.5236)
        url = dest.google_maps_url()
        self.assertTrue(url.startswith("https://www.google.com/maps/search/?api=1&query="))
        self.assertIn("Alun-Alun", url)

    def test_google_maps_url_fallback_to_name(self):
        dest = self._destination("Alun-Alun Kota Batu", latitude=None, longitude=None)
        url = dest.google_maps_url()
        self.assertIn("https://www.google.com/maps/search/?api=1&query=", url)
        self.assertIn("Kota", url)

    def test_google_maps_query_manual_overrides(self):
        dest = self._destination(
            "Batu Love Garden",
            google_maps_query="Batu Love Garden BALOGA, Kota Batu",
        )
        url = dest.google_maps_url()
        self.assertIn("Batu%20Love%20Garden%20BALOGA", url)

    # ---------------- UPDATE HARGA (fallback) ----------------

    def test_price_update_fallback_keeps_last_price(self):
        dest = self._destination(
            "Manual",
            ticket_type="fixed",
            ticket_price_weekday=25000,
        )
        result = PriceSourceService.update_destination(dest)
        dest.refresh_from_db()

        self.assertEqual(result["status"], "unavailable")
        self.assertEqual(dest.ticket_price_weekday, 25000)
        self.assertEqual(dest.price_source, "")  # tidak ditimpa

    # ---------------- EXCEL ----------------

    def test_excel_build_returns_valid_workbook(self):
        dest = self._destination(
            "Destinasi Excel",
            ticket_type="fixed",
            ticket_price_weekday=20000,
            parking_cost=5000,
            opening_time=time(8, 0),
            closing_time=time(17, 0),
        )
        params = {
            "duration_days": 1,
            "start_date": "2026-08-27",
            "traveler_count": 4,
            "transportation": "car",
        }
        content = ItineraryExcelService.build(
            params, destination_ids=[dest.id]
        )

        wb = load_workbook(BytesIO(content))
        self.assertEqual(
            wb.sheetnames,
            ["Rundown", "Ringkasan Biaya", "Destinasi", "Detail Harga"],
        )

        rundown = wb["Rundown"]
        # Judul ada di baris pertama.
        self.assertIn("RUNDOWN", rundown.cell(row=1, column=1).value)

        destinasi = wb["Destinasi"]
        self.assertEqual(destinasi.cell(row=2, column=2).value, "Destinasi Excel")
        self.assertEqual(destinasi.cell(row=2, column=8).value, "Buka Google Maps")
        self.assertTrue(destinasi.cell(row=2, column=8).hyperlink is not None)


class TouristDestinationFormTest(TestCase):
    """Checkbox "Gratis" & validasi harga di form."""

    def setUp(self):
        self.district = District.objects.create(code="T3", name="Kecamatan 3")
        self.village = Village.objects.create(
            code="T3V1", name="Desa 3", district=self.district
        )

    def _data(self, **kwargs):
        data = {
            "name": "Destinasi Form",
            "village": self.village.id,
            "is_active": "True",
            "ticket_type": "fixed",
        }
        data.update(kwargs)
        return data

    def test_free_checkbox_bypasses_price_requirement(self):
        form = TouristDestinationForm(
            data=self._data(is_free="on", ticket_type="fixed")
        )
        self.assertTrue(form.is_valid(), form.errors)
        dest = form.save()
        self.assertTrue(dest.is_free)
        self.assertEqual(dest.ticket_type, "unknown")
        self.assertIsNone(dest.ticket_price_weekday)

    def test_fixed_without_price_is_invalid(self):
        form = TouristDestinationForm(
            data=self._data(is_free="", ticket_type="fixed")
        )
        self.assertFalse(form.is_valid())
        self.assertIn("ticket_price_weekday", form.errors)

    def test_free_without_ticket_type_is_valid(self):
        # Saat "Gratis" dicentang, field "Tipe Harga" dinonaktifkan (tidak
        # dikirim). Form harus tetap valid dan tipe disimpan sebagai unknown.
        data = self._data(is_free="on")
        data.pop("ticket_type", None)
        form = TouristDestinationForm(data=data)
        self.assertTrue(form.is_valid(), form.errors)
        dest = form.save()
        self.assertTrue(dest.is_free)
        self.assertEqual(dest.ticket_type, "unknown")


class RelationalWahanaBundleTest(TestCase):
    """Model relational Wahana & TicketBundle: validasi, sync form, persistensi."""

    def setUp(self):
        self.district = District.objects.create(code="R1", name="Kecamatan R")
        self.village = Village.objects.create(
            code="R1V1", name="Desa R", district=self.district
        )

    def _destination(self, name, **kwargs):
        defaults = {"name": name, "village": self.village}
        defaults.update(kwargs)
        return TouristDestination.objects.create(**defaults)

    def _form_data(self, dest, wahana_data="[]", bundle_data="[]", **extra):
        data = {
            "name": dest.name,
            "village": self.village.id,
            "is_active": "True",
            "ticket_type": "unknown",
            "wahana_data": wahana_data,
            "bundle_data": bundle_data,
        }
        data.update(extra)
        return data

    # ---------------- Validasi pricing_type (model.clean) ----------------

    def test_included_in_htm_must_not_have_price(self):
        dest = self._destination("D1")
        w = Wahana(
            destination=dest,
            name="Taman Anak",
            pricing_type="INCLUDED_IN_HTM",
            price=5000,
        )
        with self.assertRaises(ValidationError):
            w.full_clean()

    def test_paid_requires_price(self):
        dest = self._destination("D1")
        w = Wahana(destination=dest, name="Flying Fox", pricing_type="INDEPENDENT_PRICE")
        with self.assertRaises(ValidationError):
            w.full_clean()

    def test_price_unknown_must_not_have_price(self):
        dest = self._destination("D1")
        w = Wahana(
            destination=dest, name="ATV", pricing_type="PRICE_UNKNOWN", price=1000
        )
        with self.assertRaises(ValidationError):
            w.full_clean()

    # ---------------- Sync form -> relational ----------------

    def test_form_save_creates_wahana_and_bundle(self):
        dest = self._destination("Batu Ekonomis Park")
        wahana_data = json.dumps([
            {"id": None, "name": "Flying Fox", "pricing_type": "INDEPENDENT_PRICE", "price": 15000},
            {"id": None, "name": "Taman Bermain Anak", "pricing_type": "INCLUDED_IN_HTM", "price": None},
        ])
        bundle_data = json.dumps([
            {
                "id": None,
                "name": "Paket Hemat",
                "price": 50000,
                "includes_entry_ticket": True,
                "ride_names": ["Flying Fox"],
            }
        ])
        form = TouristDestinationForm(
            instance=dest,
            data=self._form_data(dest, wahana_data, bundle_data),
        )
        self.assertTrue(form.is_valid(), form.errors)
        form.save()

        dest.refresh_from_db()
        wahanas = list(dest.wahanas.order_by("name"))
        self.assertEqual(len(wahanas), 2)
        self.assertEqual(wahanas[0].pricing_type, "INDEPENDENT_PRICE")
        self.assertEqual(wahanas[1].pricing_type, "INCLUDED_IN_HTM")
        self.assertIsNone(wahanas[1].price)

        bundles = list(dest.bundles.all())
        self.assertEqual(len(bundles), 1)
        self.assertEqual(bundles[0].price, 50000)
        self.assertTrue(bundles[0].includes_entry_ticket)
        self.assertEqual(
            [w.name for w in bundles[0].wahanas.all()], ["Flying Fox"]
        )

    # ---------------- Persistensi (data tidak hilang saat edit) ----------------

    def test_reopen_form_keeps_wahana_and_bundle(self):
        dest = self._destination("Batu Ekonomis Park")
        ff = Wahana.objects.create(
            destination=dest, name="Flying Fox", pricing_type="INDEPENDENT_PRICE", price=15000
        )
        bundle = TicketBundle.objects.create(
            destination=dest, name="Paket Hemat", price=50000,
            includes_entry_ticket=True,
        )
        bundle.wahanas.add(ff)

        # Buka form (mode edit): data existing harus muncul sebagai JSON awal.
        form = TouristDestinationForm(instance=dest)
        initial_wahana = json.loads(form.fields["wahana_data"].initial)
        initial_bundle = json.loads(form.fields["bundle_data"].initial)
        self.assertEqual(len(initial_wahana), 1)
        self.assertEqual(initial_wahana[0]["name"], "Flying Fox")
        self.assertEqual(initial_bundle[0]["name"], "Paket Hemat")
        self.assertEqual(initial_bundle[0]["ride_names"], ["Flying Fox"])

    def test_repeated_save_does_not_duplicate_or_lose_data(self):
        dest = self._destination("Batu Ekonomis Park")
        wahana_data = json.dumps([
            {"id": None, "name": "Flying Fox", "pricing_type": "INDEPENDENT_PRICE", "price": 15000},
            {"id": None, "name": "ATV", "pricing_type": "INDEPENDENT_PRICE", "price": 50000},
        ])
        bundle_data = json.dumps([
            {"id": None, "name": "Paket Hemat", "price": 50000,
             "includes_entry_ticket": False, "ride_names": ["Flying Fox"]},
        ])

        form = TouristDestinationForm(
            instance=dest,
            data=self._form_data(dest, wahana_data, bundle_data),
        )
        self.assertTrue(form.is_valid(), form.errors)
        form.save()

        # Simpan ulang (mis. edit deskripsi saja), data wahana/bundle harus
        # tetap utuh & tidak terduplikasi.
        dest.description = "diedit"
        dest.save()
        form2 = TouristDestinationForm(
            instance=dest,
            data=self._form_data(
                dest, wahana_data, bundle_data, description="diedit"
            ),
        )
        self.assertTrue(form2.is_valid(), form2.errors)
        form2.save()

        self.assertEqual(dest.wahanas.count(), 2)
        self.assertEqual(dest.bundles.count(), 1)
        self.assertEqual(dest.bundles.first().wahanas.count(), 1)

    # ---------------- Update harga (edit record, bukan buat baru) ----------------

    def test_edit_price_updates_existing_record(self):
        dest = self._destination("Batu Ekonomis Park")
        ff = Wahana.objects.create(
            destination=dest, name="Flying Fox", pricing_type="INDEPENDENT_PRICE", price=15000
        )
        wahana_data = json.dumps([
            {"id": ff.id, "name": "Flying Fox", "pricing_type": "INDEPENDENT_PRICE", "price": 20000},
        ])
        form = TouristDestinationForm(
            instance=dest,
            data=self._form_data(dest, wahana_data, "[]"),
        )
        self.assertTrue(form.is_valid(), form.errors)
        form.save()

        self.assertEqual(dest.wahanas.count(), 1)
        ff.refresh_from_db()
        self.assertEqual(ff.price, 20000)

    # ---------------- Delete bundle / wahana ----------------

    def test_delete_bundle_keeps_wahana(self):
        dest = self._destination("Batu Ekonomis Park")
        ff = Wahana.objects.create(
            destination=dest, name="Flying Fox", pricing_type="INDEPENDENT_PRICE", price=15000
        )
        bundle = TicketBundle.objects.create(
            destination=dest, name="Paket Hemat", price=50000
        )
        bundle.wahanas.add(ff)

        form = TouristDestinationForm(
            instance=dest,
            data=self._form_data(
                dest,
                json.dumps([{"id": ff.id, "name": "Flying Fox",
                             "pricing_type": "INDEPENDENT_PRICE", "price": 15000}]),
                "[]",
            ),
        )
        self.assertTrue(form.is_valid(), form.errors)
        form.save()

        self.assertEqual(dest.bundles.count(), 0)
        self.assertEqual(dest.wahanas.count(), 1)

    def test_delete_wahana_cleans_bundle_relation(self):
        dest = self._destination("Batu Ekonomis Park")
        ff = Wahana.objects.create(
            destination=dest, name="Flying Fox", pricing_type="INDEPENDENT_PRICE", price=15000
        )
        atv = Wahana.objects.create(
            destination=dest, name="ATV", pricing_type="INDEPENDENT_PRICE", price=50000
        )
        bundle = TicketBundle.objects.create(
            destination=dest, name="Paket Hemat", price=50000
        )
        bundle.wahanas.add(ff, atv)

        # Hapus "Flying Fox" dari form (hanya ATV tersisa).
        form = TouristDestinationForm(
            instance=dest,
            data=self._form_data(
                dest,
                json.dumps([{"id": atv.id, "name": "ATV",
                             "pricing_type": "INDEPENDENT_PRICE", "price": 50000}]),
                json.dumps([{"id": bundle.id, "name": "Paket Hemat", "price": 50000,
                             "includes_entry_ticket": False, "ride_names": ["ATV"]}]),
            ),
        )
        self.assertTrue(form.is_valid(), form.errors)
        form.save()

        self.assertEqual(dest.wahanas.count(), 1)
        bundle.refresh_from_db()
        # Relasi ke wahana yang dihapus otomatis bersih (tidak orphan).
        self.assertEqual([w.name for w in bundle.wahanas.all()], ["ATV"])


class DestinationFormViewTest(TestCase):
    """Integrasi view: POST menyimpan wahana/bundle, GET menampilkan lagi."""

    def setUp(self):
        self.staff = User.objects.create_user(
            username="staffform", password="x", is_staff=True
        )
        self.district = District.objects.create(code="V1", name="Kecamatan V")
        self.village = Village.objects.create(
            code="V1V1", name="Desa V", district=self.district
        )
        self.dest = TouristDestination.objects.create(
            name="Destinasi View", village=self.village, ticket_type="unknown"
        )

    def _post(self, wahana_data, bundle_data):
        return self.client.post(
            reverse("gis:manage-destination-update", args=[self.dest.pk]),
            {
                "name": "Destinasi View",
                "village": self.village.id,
                "is_active": "True",
                "ticket_type": "unknown",
                "wahana_data": wahana_data,
                "bundle_data": bundle_data,
            },
        )

    def test_post_saves_wahana_and_bundle(self):
        self.client.force_login(self.staff)
        response = self._post(
            json.dumps([
                {"id": None, "name": "Flying Fox", "pricing_type": "INDEPENDENT_PRICE", "price": 15000},
                {"id": None, "name": "Taman Anak", "pricing_type": "INCLUDED_IN_HTM", "price": None},
            ]),
            json.dumps([
                {"id": None, "name": "Paket Hemat", "price": 50000,
                 "includes_entry_ticket": True, "ride_names": ["Flying Fox"]},
            ]),
        )
        self.assertEqual(response.status_code, 302)

        self.dest.refresh_from_db()
        self.assertEqual(self.dest.wahanas.count(), 2)
        self.assertEqual(self.dest.bundles.count(), 1)
        self.assertEqual(
            list(self.dest.bundles.first().wahanas.values_list("name", flat=True)),
            ["Flying Fox"],
        )

    def test_get_renders_existing_wahana_json(self):
        ff = Wahana.objects.create(
            destination=self.dest, name="Flying Fox", pricing_type="INDEPENDENT_PRICE", price=15000
        )
        self.client.force_login(self.staff)
        response = self.client.get(
            reverse("gis:manage-destination-update", args=[self.dest.pk])
        )
        self.assertEqual(response.status_code, 200)
        # JSON wahana existing harus tampil di hidden input (repeater).
        # Nilai di-escape HTML (&quot;) oleh Django, tapi browser memulihkan
        # karakter `"` saat membaca el.value, sehingga JSON.parse tetap valid.
        self.assertContains(response, "Flying Fox")
        self.assertContains(response, "id_wahana_data")
        # Pastikan bukan JSON kosong (pricing_type tersimpan).
        self.assertContains(response, "INDEPENDENT_PRICE")


class ItineraryViewTest(TestCase):
    """Halaman itinerary & unduh Excel (visitor login boleh akses)."""

    def setUp(self):
        self.user = User.objects.create_user(username="visitor", password="x")
        self.district = District.objects.create(code="T2", name="Kecamatan 2")
        self.village = Village.objects.create(
            code="T2V1",
            name="Desa 2",
            district=self.district,
            latitude=-7.8700,
            longitude=112.5200,
        )
        TouristDestination.objects.create(
            name="Destinasi View",
            village=self.village,
            latitude=-7.8700,
            longitude=112.5200,
            ticket_type="fixed",
            ticket_price_weekday=15000,
            estimated_duration_minutes=120,
        )

    def test_itinerary_page_get(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse("gis:itinerary"))
        self.assertEqual(response.status_code, 200)

    def test_itinerary_page_post_renders_rundown(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("gis:itinerary"),
            {"duration_days": 1, "traveler_count": 2, "transportation": "car"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Rundown Perjalanan")

    def test_itinerary_excel_download(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("gis:itinerary-excel"),
            {"duration_days": 1, "traveler_count": 2, "transportation": "car"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


class ManagePaginationTest(TestCase):
    """Setelah update, admin kembali ke halaman & filter sebelumnya."""

    def setUp(self):
        self.staff = User.objects.create_user(
            username="staff2", password="x", is_staff=True
        )
        self.district = District.objects.create(code="T4", name="Kecamatan 4")
        self.village = Village.objects.create(
            code="T4V1", name="Desa 4", district=self.district
        )
        self.dest = TouristDestination.objects.create(
            name="Destinasi Paging", village=self.village
        )

    def test_update_preserves_next(self):
        self.client.force_login(self.staff)
        url = reverse("gis:manage-destination-update", args=[self.dest.pk])
        response = self.client.post(
            url,
            {
                "name": "Destinasi Paging Baru",
                "is_active": "True",
                "next": "/gis/manage/destinations/?page=5&q=foo",
            },
        )
        self.assertRedirects(
            response,
            "/gis/manage/destinations/?page=5&q=foo",
            fetch_redirect_response=False,
        )

    def test_update_price_preserves_next(self):
        self.client.force_login(self.staff)
        url = reverse("gis:manage-destination-update-price", args=[self.dest.pk])
        response = self.client.post(
            url,
            {"next": "/gis/manage/destinations/?page=3"},
        )
        self.assertRedirects(
            response,
            "/gis/manage/destinations/?page=3",
            fetch_redirect_response=False,
        )
